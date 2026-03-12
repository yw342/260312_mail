# -*- coding: utf-8 -*-
"""
재고 웹 입력·확인 및 부족 시 메일 발송
- 보내는 사람: byw004422@gmail.com
- 받는 사람: HTML 페이지 담당자 이메일 입력란
"""
import os
import json
from datetime import datetime, timedelta, timezone
from pathlib import Path

from flask import Flask, render_template, request, jsonify, redirect, url_for, session

try:
    from dotenv import load_dotenv
    load_dotenv(Path(__file__).resolve().parent / ".env")
except ImportError:
    pass

import inventory_alert as alert

app = Flask(__name__)
app.secret_key = os.environ.get("FLASK_SECRET_KEY", "inventory-app-secret-change-in-production")
# 페이지 접속: 3082=일반, 3082!@=관리자
PAGE_PASSWORD = os.environ.get("PAGE_PASSWORD", "3082")
ADMIN_PASSWORD = os.environ.get("ADMIN_PASSWORD", "3082!@")
BASE_DIR = Path(__file__).resolve().parent
EXCEL_PATH = Path(os.environ.get("INVENTORY_EXCEL_PATH", str(BASE_DIR / "domino_inventory_training.xlsx")))
# 이메일 발송 이력: Supabase 사용 시 DB, 미설정 시 파일
SUPABASE_URL = os.environ.get("SUPABASE_URL", "").strip()
SUPABASE_KEY = os.environ.get("SUPABASE_SERVICE_ROLE_KEY", os.environ.get("SUPABASE_KEY", "")).strip()
USE_SUPABASE = bool(SUPABASE_URL and SUPABASE_KEY)

if not USE_SUPABASE:
    if os.environ.get("VERCEL"):
        EMAIL_HISTORY_PATH = Path("/tmp/email_history.json")
    else:
        EMAIL_HISTORY_PATH = BASE_DIR / "email_history.json"
EMAIL_THRESHOLD_HOURS = 1

_supabase_client = None

def _get_supabase():
    if not USE_SUPABASE:
        return None
    global _supabase_client
    if _supabase_client is None:
        from supabase import create_client
        _supabase_client = create_client(SUPABASE_URL, SUPABASE_KEY)
    return _supabase_client


def _load_email_history_file():
    if USE_SUPABASE:
        return []
    path = EMAIL_HISTORY_PATH
    if not path.is_file():
        return []
    try:
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return []


def _save_email_history_file(records):
    if USE_SUPABASE:
        return
    with open(EMAIL_HISTORY_PATH, "w", encoding="utf-8") as f:
        json.dump(records[-500:], f, ensure_ascii=False, indent=2)


def load_email_history():
    """이메일 발송 이력 로드 (Supabase 또는 파일)."""
    if USE_SUPABASE:
        try:
            sb = _get_supabase()
            r = sb.table("email_send_history").select("id,sent_at,to_email,item_codes,item_names").order("sent_at", desc=True).limit(500).execute()
            rows = r.data or []
            return [
                {"id": x.get("id"), "sent_at": x.get("sent_at"), "to": x.get("to_email"), "item_codes": x.get("item_codes") or [], "item_names": x.get("item_names") or []}
                for x in rows
            ]
        except Exception:
            return []
    return _load_email_history_file()


def get_item_codes_sent_within_hours(hours=1):
    """지정 시간(기본 1시간) 내 발송된 품목의 품목코드 set 반환."""
    records = load_email_history()
    cutoff = datetime.now(timezone.utc) - timedelta(hours=hours)
    codes = set()
    for r in records:
        try:
            sent_at = datetime.fromisoformat(r.get("sent_at", "").replace("Z", "+00:00"))
            if sent_at.tzinfo is None:
                sent_at = sent_at.replace(tzinfo=timezone.utc)
        except Exception:
            continue
        if sent_at >= cutoff:
            for code in r.get("item_codes") or []:
                if code:
                    codes.add(str(code).strip())
    return codes


def append_email_record(to_email, items):
    """발송 이력에 한 건 추가 (Supabase 또는 파일)."""
    item_codes = [str(i.get("품목코드") or "").strip() for i in items if i.get("품목코드")]
    item_names = [str(i.get("재료명") or "").strip() for i in items]
    sent_at = datetime.now(timezone.utc).isoformat().replace("+00:00", "Z")

    if USE_SUPABASE:
        try:
            sb = _get_supabase()
            sb.table("email_send_history").insert({
                "sent_at": sent_at,
                "to_email": to_email,
                "item_codes": item_codes,
                "item_names": item_names,
            }).execute()
        except Exception:
            pass
        return
    records = _load_email_history_file()
    records.append({"sent_at": sent_at, "to": to_email, "item_codes": item_codes, "item_names": item_names})
    _save_email_history_file(records)


def get_email_history_for_display(limit=50):
    """하단 이력 표시용. 최근 limit건. 각 행에 id 포함 (삭제용)."""
    records = load_email_history()
    out = []
    iter_records = records[:limit] if USE_SUPABASE else list(reversed(records[-limit:]))
    for i, r in enumerate(iter_records):
        try:
            sent_at = r.get("sent_at", "")
            if sent_at:
                dt = datetime.fromisoformat(sent_at.replace("Z", "+00:00"))
                sent_at = dt.strftime("%Y-%m-%d %H:%M")
        except Exception:
            pass
        record_id = r.get("id") if USE_SUPABASE else ("file_" + str(i))
        out.append({
            "id": record_id,
            "sent_at": sent_at,
            "to": r.get("to") or r.get("to_email", ""),
            "item_names": r.get("item_names") or [],
        })
    return out


def _num_display(val):
    """소수 부분이 0이면 정수로, 아니면 그대로 반환 (화면 표시용)."""
    if val is None or val == "":
        return val
    if isinstance(val, int):
        return val
    if isinstance(val, float):
        try:
            if val == int(val):
                return int(val)
        except (ValueError, OverflowError):
            pass
        return val
    return val


def get_inventory_from_supabase():
    """Supabase inventory 테이블에서 재고 목록 로드."""
    try:
        sb = _get_supabase()
        r = sb.table("inventory").select("*").order("row_order").execute()
        rows = r.data or []
    except Exception:
        return []
    sent_within_hour = get_item_codes_sent_within_hours(EMAIL_THRESHOLD_HOURS)
    result = []
    for i, x in enumerate(rows):
        code = (x.get("item_code") or "").strip()
        current = x.get("current_stock")
        safety = x.get("safety_stock")
        try:
            current = float(current) if current is not None else 0
            safety = float(safety) if safety is not None else 0
        except (TypeError, ValueError):
            current, safety = 0, 0
        status = "발주 필요" if current < safety else "정상"
        result.append({
            "id": str(x.get("id", "")),
            "row": i + 1,
            "품목코드": code,
            "재료명": (x.get("item_name") or "").strip(),
            "규격": _num_display(x.get("spec")),
            "단위": (x.get("unit") or "").strip(),
            "현재재고": _num_display(current),
            "안전재고": _num_display(safety),
            "상태": status,
            "거래처이메일": (x.get("supplier_email") or "").strip(),
            "이메일_발송여부": "1시간 내 발송" if (code and code in sent_within_hour) else "-",
        })
    return result


def get_inventory_list():
    """재고 목록 (Supabase 우선, 없으면 엑셀)."""
    if USE_SUPABASE:
        return get_inventory_from_supabase()
    data_rows, col = alert.load_inventory(EXCEL_PATH)
    if not col:
        return []
    idx_code = col.get("품목코드", -1)
    idx_name = col.get("재료명", -1)
    idx_spec = col.get("규격", -1)
    idx_unit = col.get("단위", -1)
    idx_current = col.get("현재재고", col.get("현재 재고", -1))
    idx_safety = col.get("안전재고", col.get("안전 재고", -1))
    idx_status = col.get("상태", -1)
    idx_email = col.get("거래처이메일", col.get("이메일", -1))

    def v(row, i):
        if i < 0 or i >= len(row):
            return ""
        x = row[i]
        if x is None:
            return ""
        return str(x).strip() if not isinstance(x, (int, float)) else x

    sent_within_hour = get_item_codes_sent_within_hours(EMAIL_THRESHOLD_HOURS)
    result = []
    for i, row in enumerate(data_rows):
        excel_row = i + 2
        raw_spec = v(row, idx_spec)
        raw_current = v(row, idx_current)
        raw_safety = v(row, idx_safety)
        code = v(row, idx_code)
        result.append({
            "id": None,
            "row": excel_row,
            "품목코드": code,
            "재료명": v(row, idx_name),
            "규격": _num_display(raw_spec) if isinstance(raw_spec, (int, float)) else raw_spec,
            "단위": v(row, idx_unit),
            "현재재고": _num_display(raw_current) if isinstance(raw_current, (int, float)) else raw_current,
            "안전재고": _num_display(raw_safety) if isinstance(raw_safety, (int, float)) else raw_safety,
            "상태": v(row, idx_status),
            "거래처이메일": v(row, idx_email),
            "이메일_발송여부": "1시간 내 발송" if (code and str(code).strip() in sent_within_hour) else "-",
        })
    return result


def update_inventory_supabase(updates):
    """Supabase inventory 테이블 업데이트 (현재재고, 거래처이메일)."""
    sb = _get_supabase()
    for u in updates:
        uid = u.get("id")
        if not uid:
            continue
        payload = {}
        if "현재재고" in u:
            try:
                payload["current_stock"] = float(u["현재재고"])
            except (TypeError, ValueError):
                payload["current_stock"] = u["현재재고"]
        if "거래처이메일" in u:
            payload["supplier_email"] = (u.get("거래처이메일") or "").strip()
        if payload:
            sb.table("inventory").update(payload).eq("id", uid).execute()


def update_excel_current_stock(updates):
    """updates: [ {"row": 2, "현재재고": 120, "거래처이메일": "..."}, ... ] → 엑셀 해당 행 수정."""
    import openpyxl
    try:
        wb = openpyxl.load_workbook(EXCEL_PATH, data_only=False)
    except Exception:
        return False
    if "Inventory" not in wb.sheetnames:
        wb.close()
        return False
    ws = wb["Inventory"]
    header = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    col_current = col_email = None
    for i, h in enumerate(header):
        if h and ("현재재고" in str(h) or "현재 재고" in str(h)):
            col_current = i + 1
        if h and ("거래처이메일" in str(h) or "이메일" == str(h).strip()):
            col_email = i + 1
    for u in updates:
        row = u.get("row")
        if row is None:
            continue
        if col_current is not None and "현재재고" in u:
            val = u["현재재고"]
            try:
                n = float(val)
                ws.cell(row=int(row), column=col_current, value=n)
            except (TypeError, ValueError):
                ws.cell(row=int(row), column=col_current, value=val)
        if col_email is not None and "거래처이메일" in u:
            ws.cell(row=int(row), column=col_email, value=(u.get("거래처이메일") or "").strip())
    wb.save(EXCEL_PATH)
    wb.close()
    return True


def get_dashboard(items):
    """재고 목록에서 대시보드용 집계 (총 품목, 정상, 부족, 부족 품목명)."""
    total = len(items)
    low_stock_names = []
    for it in items:
        try:
            current = float(it.get("현재재고") or 0)
            safety = float(it.get("안전재고") or 0)
        except (TypeError, ValueError):
            current, safety = 0, 0
        if current < safety:
            low_stock_names.append(it.get("재료명") or "-")
    low_count = len(low_stock_names)
    return {
        "total": total,
        "normal_count": total - low_count,
        "low_stock_count": low_count,
        "low_stock_names": low_stock_names,
    }


@app.before_request
def require_login():
    if request.endpoint in ("login", "logout", "static"):
        return None
    if session.get("authenticated"):
        return None
    return redirect(url_for("login"))


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        password = (request.form.get("password") or (request.get_json(silent=True) or {}).get("password") or "").strip()
        if password == ADMIN_PASSWORD:
            session["authenticated"] = True
            session["role"] = "admin"
            return redirect(url_for("index"))
        if password == PAGE_PASSWORD:
            session["authenticated"] = True
            session["role"] = "user"
            return redirect(url_for("index"))
        return render_template("login.html", error="비밀번호가 올바르지 않습니다.")
    return render_template("login.html", error=None)


@app.route("/logout")
def logout():
    session.pop("authenticated", None)
    session.pop("role", None)
    return redirect(url_for("login"))


def _is_admin():
    return session.get("role") == "admin"


@app.route("/")
def index():
    items = get_inventory_list()
    dashboard = get_dashboard(items)
    email_history = get_email_history_for_display(50)
    return render_template("index.html", items=items, dashboard=dashboard, email_history=email_history, is_admin=_is_admin())


def get_low_stock_from_items(items):
    """웹용 재고 목록에서 현재재고 < 안전재고 인 품목만 추출 (alert.build_email_body 형식)."""
    low = []
    for it in items:
        try:
            current = float(it.get("현재재고") or 0)
            safety = float(it.get("안전재고") or 0)
        except (TypeError, ValueError):
            current, safety = 0, 0
        if current >= safety:
            continue
        name = (it.get("재료명") or "").strip() or "-"
        unit = (it.get("단위") or "").strip()
        order_qty = max(0, safety - current)
        msg = f"{name} 재고 부족 - 현재 {int(current)}{unit}, 안전재고 {int(safety)}{unit}, 권장발주 {order_qty}{unit}"
        low.append({
            "품목코드": (it.get("품목코드") or "").strip(),
            "재료명": name,
            "단위": unit,
            "현재재고": current,
            "안전재고": safety,
            "발주권장수량": order_qty,
            "담당자알림메시지": msg,
            "거래처이메일": (it.get("거래처이메일") or "").strip(),
        })
    return low


def _group_low_stock_by_email(low_stock):
    """재고 부족 품목을 거래처이메일별로 묶음. 유효한 이메일만."""
    by_email = {}
    for x in low_stock:
        email = (x.get("거래처이메일") or "").strip()
        if not email or "@" not in email:
            continue
        if email not in by_email:
            by_email[email] = []
        by_email[email].append(x)
    return by_email


def check_and_send_alert():
    """재고 부족 확인 후 각 품목의 거래처이메일로 메일 발송(거래처별 1통). 1시간 내 발송 품목 제외."""
    if not alert.SENDER_PASSWORD:
        return False, 0, "INVENTORY_SENDER_PASSWORD 환경 변수(또는 .env) 설정 필요"
    try:
        items = get_inventory_list()
        if not items:
            return False, 0, "재고 데이터가 없습니다." if USE_SUPABASE else "Inventory 시트 헤더 없음"
        low_stock = get_low_stock_from_items(items)
        if not low_stock:
            return False, 0, None
        sent_recently = get_item_codes_sent_within_hours(EMAIL_THRESHOLD_HOURS)
        low_stock_to_send = [x for x in low_stock if (x.get("품목코드") or "").strip() not in sent_recently]
        if not low_stock_to_send:
            return False, 0, "1시간 내 발송된 품목만 있어 이번에는 발송하지 않았습니다."
        by_email = _group_low_stock_by_email(low_stock_to_send)
        if not by_email:
            return False, 0, "발송할 품목에 유효한 거래처 이메일이 없습니다."
        total_sent = 0
        for to_email, items_for_email in by_email.items():
            subject, body = alert.build_email_body(
                to_email, items_for_email, alert.SENDER_EMAIL, all_to_one=True
            )
            alert.send_mail(
                to_email, subject, body,
                alert.SENDER_EMAIL, alert.SENDER_PASSWORD
            )
            append_email_record(to_email, items_for_email)
            total_sent += len(items_for_email)
        return True, total_sent, None
    except Exception as e:
        return False, 0, str(e)


@app.route("/save", methods=["POST"])
def save():
    data = request.get_json(force=True, silent=True) or request.form
    if not data:
        return jsonify({"ok": False, "message": "데이터 없음"}), 400
    updates = data.get("updates") if isinstance(data, dict) else []
    if not updates:
        return jsonify({"ok": False, "message": "updates 배열 필요"}), 400
    if not _is_admin():
        return jsonify({"ok": False, "message": "재고·거래처 이메일 수정은 관리자만 가능합니다."}), 403
    try:
        if USE_SUPABASE:
            update_inventory_supabase(updates)
        else:
            update_excel_current_stock(updates)
        email_sent, alert_count, alert_error = check_and_send_alert()
        res = {"ok": True, "email_sent": email_sent, "alert_count": alert_count}
        if alert_error and email_sent is False and alert_count == 0:
            res["alert_error"] = alert_error
        return jsonify(res)
    except Exception as e:
        return jsonify({"ok": False, "message": str(e)}), 500


@app.route("/send-alert", methods=["POST"])
def send_alert():
    """재고 부족 확인 후 각 품목의 거래처이메일로 메일 발송."""
    email_sent, alert_count, error = check_and_send_alert()
    if error and not email_sent:
        return jsonify({"ok": False, "message": error}), 400
    if not email_sent:
        return jsonify({"ok": True, "sent": 0, "message": "재고 부족 품목 없음"})
    return jsonify({"ok": True, "sent": 1, "count": alert_count, "message": "거래처 이메일로 발송 완료"})


def delete_email_history_record(record_id):
    """발송 이력 한 건 삭제. Supabase는 id(uuid), 파일은 file_0 형식."""
    if not record_id:
        return False, "id 없음"
    if USE_SUPABASE:
        try:
            sb = _get_supabase()
            sb.table("email_send_history").delete().eq("id", record_id).execute()
            return True, None
        except Exception as e:
            return False, str(e)
    # 파일 모드: file_0, file_1 ... = 표시 순서(0=최신)
    if isinstance(record_id, str) and record_id.startswith("file_"):
        try:
            i = int(record_id.split("_")[1])
            records = _load_email_history_file()
            # 표시 순서 i = 리스트에서 len(records)-1-i 번째 (최신이 마지막)
            idx = len(records) - 1 - i
            if idx < 0 or idx >= len(records):
                return False, "범위 오류"
            records.pop(idx)
            _save_email_history_file(records)
            return True, None
        except (ValueError, IndexError) as e:
            return False, str(e)
    return False, "잘못된 id"


@app.route("/delete-email-history", methods=["POST"])
def delete_email_history():
    """이메일 발송 이력 한 건 삭제. 관리자만 가능."""
    if not _is_admin():
        return jsonify({"ok": False, "message": "이력 삭제는 관리자만 가능합니다."}), 403
    data = request.get_json(force=True, silent=True) or {}
    record_id = (data.get("id") or "").strip() or data.get("id")
    if not record_id:
        return jsonify({"ok": False, "message": "id 필요"}), 400
    ok, err = delete_email_history_record(record_id)
    if not ok:
        return jsonify({"ok": False, "message": err or "삭제 실패"}), 400
    return jsonify({"ok": True})


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=True)
