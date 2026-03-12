# -*- coding: utf-8 -*-
"""
재고 파악 후 부족 시 담당 기업 직원에게 메일을 보내는 자동화 스크립트
- 엑셀 'Inventory' 시트에서 현재재고 < 안전재고인 품목을 찾습니다.
- 부족 품목이 있으면 해당 거래처 담당자(거래처이메일)에게 발주 요청 메일을 발송합니다.
- 발신 주소: byw004422@gmail.com (설정에서 변경 가능)
"""

import os
import sys
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from pathlib import Path

# .env 파일 지원 (선택)
try:
    from dotenv import load_dotenv
    load_dotenv(Path(__file__).resolve().parent / ".env")
except ImportError:
    pass

try:
    import openpyxl
except ImportError:
    print("openpyxl이 필요합니다. 다음 명령으로 설치하세요: pip install openpyxl")
    sys.exit(1)


# ---------- 설정 ----------
SENDER_EMAIL = os.environ.get("INVENTORY_SENDER_EMAIL", "byw004422@gmail.com")
# 재고 부족 알림 수신 주소 (엑셀 거래처이메일 대신 이 주소로 모두 발송)
RECIPIENT_EMAIL = os.environ.get("INVENTORY_RECIPIENT_EMAIL", "ibk6895@gmail.com")
# Gmail 앱 비밀번호 또는 계정 비밀번호 (환경변수 INVENTORY_SENDER_PASSWORD 또는 아래에 설정)
SENDER_PASSWORD = os.environ.get("INVENTORY_SENDER_PASSWORD", "")
EXCEL_FILE = Path(__file__).resolve().parent / "domino_inventory_training.xlsx"
INVENTORY_SHEET = "Inventory"
# 재고 부족 조건: 현재재고 < 안전재고 (또는 상태가 "발주 필요")
LOW_STOCK_STATUS = "발주 필요"


def load_inventory(excel_path: Path):
    """엑셀에서 Inventory 시트를 읽어 데이터 행 목록과 헤더 매핑을 반환합니다."""
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    if INVENTORY_SHEET not in wb.sheetnames:
        raise FileNotFoundError(f"엑셀에 '{INVENTORY_SHEET}' 시트가 없습니다.")
    ws = wb[INVENTORY_SHEET]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return [], {}
    header = [str(c).strip() if c is not None else "" for c in rows[0]]
    col = {h: i for i, h in enumerate(header) if h}
    data_rows = []
    for row in rows[1:]:
        if row and any(c is not None for c in row):
            data_rows.append(row)
    return data_rows, col


def get_low_stock_items(data_rows, col):
    """현재재고 < 안전재고 인 행만 필터링합니다. (웹에서 저장한 숫자 기준으로 판단)"""
    low = []
    idx_current = col.get("현재재고", col.get("현재 재고", -1))
    idx_safety = col.get("안전재고", col.get("안전 재고", -1))
    idx_name = col.get("재료명", -1)
    idx_unit = col.get("단위", -1)
    idx_order_qty = col.get("발주권장수량", col.get("발주 권장수량", -1))
    idx_msg = col.get("담당자알림메시지", col.get("담당자 알림메시지", -1))
    idx_email = col.get("거래처이메일", col.get("이메일", -1))
    idx_supplier = col.get("거래처", -1)
    idx_contact = col.get("알림담당자", -1)

    for row in data_rows:
        def v(i):
            if i < 0 or i >= len(row):
                return None
            x = row[i]
            if x is None:
                return None
            if isinstance(x, (int, float)):
                return x
            return str(x).strip() or None

        try:
            current = float(v(idx_current) or 0)
            safety = float(v(idx_safety) or 0)
        except (TypeError, ValueError):
            current, safety = 0, 0
        if current < safety:
            low.append({
                "재료명": v(idx_name) or "-",
                "단위": v(idx_unit) or "",
                "현재재고": current,
                "안전재고": safety,
                "발주권장수량": v(idx_order_qty),
                "담당자알림메시지": v(idx_msg) or _default_alert_msg(current, safety, v(idx_name), v(idx_unit), v(idx_order_qty)),
                "거래처이메일": v(idx_email) or "",
                "거래처": v(idx_supplier) or "",
                "알림담당자": v(idx_contact) or "",
            })
    return low


def _default_alert_msg(current, safety, name, unit, order_qty):
    """담당자알림메시지가 없을 때 기본 문구."""
    name = name or "-"
    unit = unit or ""
    order_qty = order_qty if order_qty is not None else max(0, safety - current)
    return f"{name} 재고 부족 - 현재 {int(current)}{unit}, 안전재고 {int(safety)}{unit}, 권장발주 {order_qty}{unit}"


def group_by_email(low_stock_items):
    """동일한 거래처이메일별로 묶어서 { 이메일: [품목들] } 형태로 반환합니다."""
    by_email = {}
    for item in low_stock_items:
        email = (item.get("거래처이메일") or "").strip()
        if not email or "@" not in email:
            continue
        if email not in by_email:
            by_email[email] = []
        by_email[email].append(item)
    return by_email


def build_email_body(recipient_email: str, items: list, sender_email: str, all_to_one: bool = False) -> tuple:
    """메일 제목과 본문(UTF-8 텍스트)을 만듭니다."""
    if not items:
        return "", ""
    lines = []
    for it in items:
        msg = (it.get("담당자알림메시지") or "").strip()
        if msg:
            lines.append(f"· {msg}")
        else:
            name = it.get("재료명", "-")
            order_qty = it.get("발주권장수량", "")
            unit = it.get("단위", "")
            lines.append(f"· {name} - 권장 발주: {order_qty}{unit}")
    item_list = "\n".join(lines)
    if all_to_one:
        subject = "[재고 부족 알림] 발주 필요 품목 안내"
        body = f"""안녕하세요, 담당자님.

도미노피자 재고·발주 자동화 시스템에서 재고 부족 품목이 있어 안내드립니다.

【발주 필요 품목】
{item_list}

위 품목에 대해 발주 검토 부탁드립니다.
감사합니다.
"""
    else:
        supplier = (items[0].get("거래처") or "").strip() or "담당자"
        subject = f"[재고 부족 알림] {supplier} 발주 필요 품목 안내"
        body = f"""안녕하세요, {supplier} 담당자님.

도미노피자 재고·발주 자동화 시스템에서 재고 부족 품목이 있어 안내드립니다.

【발주 필요 품목】
{item_list}

위 품목에 대해 발주 검토 부탁드립니다.
감사합니다.
"""
    return subject, body


def send_mail(to_email: str, subject: str, body: str, sender_email: str, sender_password: str):
    """Gmail SMTP로 메일을 발송합니다. (보내는 사람: byw004422, 받는 사람: ibk6895)"""
    msg = MIMEMultipart("alternative")
    msg["Subject"] = subject
    msg["From"] = sender_email
    msg["To"] = to_email
    msg.attach(MIMEText(body, "plain", "utf-8"))
    with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
        server.login(sender_email, sender_password)
        server.sendmail(sender_email, [to_email], msg.as_string())


def main():
    check_only = "--check-only" in sys.argv or "-n" in sys.argv
    base_dir = Path(__file__).resolve().parent
    excel_path = os.environ.get("INVENTORY_EXCEL_PATH", str(base_dir / "domino_inventory_training.xlsx"))
    excel_path = Path(excel_path)
    if not excel_path.is_file():
        print(f"엑셀 파일을 찾을 수 없습니다: {excel_path}")
        sys.exit(1)

    password = SENDER_PASSWORD
    if not password and not check_only:
        print("발신용 Gmail 비밀번호가 설정되지 않았습니다.")
        print("  방법 1: 환경변수 INVENTORY_SENDER_PASSWORD 에 Gmail 앱 비밀번호 설정")
        print("  방법 2: .env 파일에 INVENTORY_SENDER_PASSWORD=앱비밀번호 로 설정 후 python-dotenv 사용")
        print("Gmail 앱 비밀번호 발급: Google 계정 → 보안 → 2단계 인증 → 앱 비밀번호")
        sys.exit(1)

    print("엑셀 재고 데이터 읽는 중...")
    data_rows, col = load_inventory(excel_path)
    if not col:
        print("Inventory 시트에 헤더가 없거나 비어 있습니다.")
        sys.exit(1)

    low_stock = get_low_stock_items(data_rows, col)
    if not low_stock:
        print("재고 부족 품목이 없습니다. 메일을 보내지 않습니다.")
        return

    # 수신 주소를 모두 RECIPIENT_EMAIL(ibk6895@gmail.com)로 고정
    by_email = {RECIPIENT_EMAIL: low_stock}

    if check_only:
        print(f"[확인 모드] 재고 부족 품목 {len(low_stock)}건, 다음 주소로 1통 발송됩니다: {RECIPIENT_EMAIL}")
        for it in low_stock:
            print(f"  - {it.get('담당자알림메시지', it.get('재료명', ''))}")
        print("실제 발송은 .env에 INVENTORY_SENDER_PASSWORD 설정 후 python inventory_alert.py 로 실행하세요.")
        return

    print(f"재고 부족 품목 {len(low_stock)}건 → {RECIPIENT_EMAIL} 로 메일 발송 (발신: {SENDER_EMAIL})")
    for email, items in by_email.items():
        subject, body = build_email_body(email, items, SENDER_EMAIL, all_to_one=True)
        try:
            send_mail(email, subject, body, SENDER_EMAIL, password)
            print(f"  발송 완료: {email} ({len(items)}건)")
        except Exception as e:
            print(f"  발송 실패: {email} - {e}")
            raise

    print("모든 알림 메일 발송을 완료했습니다.")


if __name__ == "__main__":
    main()
