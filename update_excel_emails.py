# -*- coding: utf-8 -*-
"""엑셀 파일 내 담당자/거래처 이메일을 지정 주소로 일괄 변경합니다."""
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("openpyxl이 필요합니다. pip install openpyxl")
    sys.exit(1)

EXCEL_PATH = Path(__file__).resolve().parent / "domino_inventory_training.xlsx"
TARGET_EMAIL = "ibk6895@gmail.com"


def main():
    if not EXCEL_PATH.is_file():
        print(f"엑셀 파일 없음: {EXCEL_PATH}")
        sys.exit(1)
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=False)
    changed = 0

    # Suppliers 시트: '이메일' 컬럼
    if "Suppliers" in wb.sheetnames:
        ws = wb["Suppliers"]
        header = [cell.value for cell in ws[1]]
        col_idx = None
        for i, h in enumerate(header):
            if h and ("이메일" in str(h) or "담당자 이메일" in str(h)):
                col_idx = i
                break
        if col_idx is not None:
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=col_idx + 1)
                if cell.value:
                    cell.value = TARGET_EMAIL
                    changed += 1

    # Inventory 시트: '거래처이메일' 컬럼
    if "Inventory" in wb.sheetnames:
        ws = wb["Inventory"]
        header = [cell.value for cell in ws[1]]
        col_idx = None
        for i, h in enumerate(header):
            if h and ("거래처이메일" in str(h) or "이메일" in str(h)):
                col_idx = i
                break
        if col_idx is not None:
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=col_idx + 1)
                cell.value = TARGET_EMAIL
                changed += 1

    wb.save(EXCEL_PATH)
    wb.close()
    print(f"엑셀 내 이메일을 {TARGET_EMAIL} 로 {changed}개 셀 수정 후 저장했습니다: {EXCEL_PATH}")


if __name__ == "__main__":
    main()
