# -*- coding: utf-8 -*-
"""
로컬에서 한 번만 실행: 엑셀 Inventory 시트 데이터를 Supabase inventory 테이블에 넣습니다.
환경변수 SUPABASE_URL, SUPABASE_SERVICE_ROLE_KEY 설정 후 실행하세요.
"""
import os
import sys
from pathlib import Path

try:
    from dotenv import load_dotenv
    load_dotenv(Path(__file__).resolve().parent / ".env")
except ImportError:
    pass

import openpyxl

BASE_DIR = Path(__file__).resolve().parent
EXCEL_PATH = BASE_DIR / "domino_inventory_training.xlsx"
SUPABASE_URL = os.environ.get("SUPABASE_URL", "").strip()
SUPABASE_KEY = os.environ.get("SUPABASE_SERVICE_ROLE_KEY", os.environ.get("SUPABASE_KEY", "")).strip()


def main():
    if not SUPABASE_URL or not SUPABASE_KEY:
        print("SUPABASE_URL, SUPABASE_SERVICE_ROLE_KEY 환경 변수를 설정하세요.")
        sys.exit(1)
    if not EXCEL_PATH.is_file():
        print(f"엑셀 파일 없음: {EXCEL_PATH}")
        sys.exit(1)
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    if "Inventory" not in wb.sheetnames:
        print("Inventory 시트가 없습니다.")
        sys.exit(1)
    ws = wb["Inventory"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        print("데이터가 없습니다.")
        sys.exit(1)
    header = [str(c).strip() if c is not None else "" for c in rows[0]]
    col = {h: i for i, h in enumerate(header) if h}
    idx_code = col.get("품목코드", -1)
    idx_name = col.get("재료명", -1)
    idx_spec = col.get("규격", -1)
    idx_unit = col.get("단위", -1)
    idx_current = col.get("현재재고", col.get("현재 재고", -1))
    idx_safety = col.get("안전재고", col.get("안전 재고", -1))
    idx_email = col.get("거래처이메일", col.get("이메일", -1))

    def v(row, i):
        if i < 0 or i >= len(row):
            return None
        x = row[i]
        if x is None:
            return None
        if isinstance(x, (int, float)):
            return x
        return str(x).strip() or None

    from supabase import create_client
    sb = create_client(SUPABASE_URL, SUPABASE_KEY)
    inserted = 0
    for i, row in enumerate(rows[1:], start=1):
        if not row or all(c is None for c in row):
            continue
        try:
            current = float(v(row, idx_current) or 0)
            safety = float(v(row, idx_safety) or 0)
        except (TypeError, ValueError):
            current, safety = 0, 0
        sb.table("inventory").insert({
            "row_order": i,
            "item_code": (v(row, idx_code) or ""),
            "item_name": (v(row, idx_name) or ""),
            "spec": (v(row, idx_spec) or ""),
            "unit": (v(row, idx_unit) or ""),
            "current_stock": current,
            "safety_stock": safety,
            "supplier_email": (v(row, idx_email) or ""),
        }).execute()
        inserted += 1
    print(f"Supabase inventory 테이블에 {inserted}건 삽입했습니다.")


if __name__ == "__main__":
    main()
